# -*- coding: utf-8 -*-
from datetime import datetime
from pathlib import Path
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule


class TrainingCostExcelReport:

    def __init__(self, report_data: Dict[str, Any]):
        self.report_data = report_data
        self.breakdown = report_data.get("breakdown", {})
        self.scenarios = report_data.get("scenarios", [])
        self.title = report_data.get("title", "AI Training Cost Report")
        self.generated_at = report_data.get("generated_at", datetime.utcnow())
        self.currency = self.breakdown.get("currency", "USD")

    def generate(self, output_path: str) -> str:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(wb)

        self._create_breakdown_sheet(wb)

        if self.scenarios:
            self._create_scenarios_sheet(wb)

        self._create_metadata_sheet(wb)

        wb.save(str(output_path))
        return str(output_path)


    def _create_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Executive Summary")

        ws.merge_cells("A1:E1")
        ws["A1"] = self.title
        ws["A1"].font = Font(size=18, bold=True, color="1F3A8A")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:E2")
        ws["A2"] = f"Generated at: {self.generated_at.strftime('%Y-%m-%d %H:%M UTC')}"

        total_cost = self.breakdown.get("total_cost", 0)

        ws["A4"] = "Total Training Cost"
        ws["B4"] = total_cost
        ws["B4"].font = Font(size=16, bold=True, color="166534")
        ws["B4"].number_format = '#,##0.00'

        ws["D4"] = "Cost per Training Hour"
        ws["E4"] = total_cost / max(self.breakdown.get("training_hours", 1), 1)
        ws["E4"].number_format = '#,##0.00'

        ws["D5"] = "Cost per 1M Tokens"
        ws["E5"] = self.breakdown.get("token_cost", 0)
        ws["E5"].number_format = '#,##0.00'

        ws["A7"] = "Cost Breakdown"
        ws["A7"].font = Font(size=14, bold=True)

        headers = ["Category", f"Cost ({self.currency})", "Percentage"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F3A8A", fill_type="solid")

        total = max(total_cost, 1)

        cost_data = [
            ("Hardware", self.breakdown.get("hardware_cost", 0)),
            ("Storage", self.breakdown.get("storage_cost", 0)),
            ("Token", self.breakdown.get("token_cost", 0)),
            ("Energy", self.breakdown.get("energy_cost", 0)),
        ]

        for row, (cat, cost) in enumerate(cost_data, start=9):
            ws.cell(row=row, column=1, value=cat)
            c = ws.cell(row=row, column=2, value=cost)
            c.number_format = '#,##0.00'
            p = ws.cell(row=row, column=3, value=cost / total)
            p.number_format = '0.0%'

        self._add_summary_bar_chart(ws, len(cost_data))
        self._add_cost_pie_chart(ws, len(cost_data))

        rule = ColorScaleRule(start_type='min', start_color='63BE7B',
                              mid_type='percentile', mid_value=50, mid_color='FFEB84',
                              end_type='max', end_color='F8696B')
        ws.conditional_formatting.add(f"B9:B{8+len(cost_data)}", rule)

    def _add_summary_bar_chart(self, ws, count):
        chart = BarChart()
        chart.title = "Cost Distribution"
        data = Reference(ws, min_col=2, min_row=8, max_row=8+count)
        cats = Reference(ws, min_col=1, min_row=9, max_row=8+count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E8")

    def _add_cost_pie_chart(self, ws, count):
        chart = PieChart()
        chart.title = "Cost Share"
        data = Reference(ws, min_col=2, min_row=9, max_row=8+count)
        labels = Reference(ws, min_col=1, min_row=9, max_row=8+count)
        chart.add_data(data)
        chart.set_categories(labels)
        ws.add_chart(chart, "E22")


    def _create_breakdown_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Detailed Breakdown")
        ws["A1"] = "Detailed Cost Breakdown"
        ws["A1"].font = Font(size=16, bold=True)

    def _create_scenarios_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Scenario Comparison")

        headers = ["Scenario", "Tags", f"Total Cost ({self.currency})",
                   "vs Baseline", "Savings", "Efficiency Score"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="166534", fill_type="solid")

        baseline = self.breakdown.get("total_cost", 0)

        for row, sc in enumerate(self.scenarios, start=3):
            cost = sc.get("total_cost", 0)
            diff = cost - baseline
            savings = baseline - cost
            score = max(0, 100 - abs((diff / baseline * 100) if baseline else 0))

            ws.cell(row=row, column=1, value=sc.get("scenario"))
            ws.cell(row=row, column=2, value=", ".join(sc.get("tags", [])))
            ws.cell(row=row, column=3, value=cost).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=diff)
            ws.cell(row=row, column=5, value=savings)
            ws.cell(row=row, column=6, value=round(score, 1))

        self._add_scenario_chart(ws, len(self.scenarios))

    def _add_scenario_chart(self, ws, count):
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Scenario Cost Comparison"
        data = Reference(ws, min_col=3, min_row=2, max_row=2+count)
        cats = Reference(ws, min_col=1, min_row=3, max_row=2+count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "H3")

    def _create_metadata_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Metadata")
        ws["A1"] = "Generated By"
        ws["B1"] = "AI Training Cost System"
        ws["A2"] = "Model Name"
        ws["B2"] = self.report_data.get("model_name", "N/A")
        ws["A3"] = "Dataset"
        ws["B3"] = self.report_data.get("dataset", "N/A")
        ws["A4"] = "GPU Type"
        ws["B4"] = self.report_data.get("gpu_type", "N/A")
